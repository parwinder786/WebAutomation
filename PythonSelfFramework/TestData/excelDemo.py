import openpyxl
#book =openpyxl.load_workbook("C:\\Users\\Owner\\Documents\\PythonDemo.xlsx")  #load the excel file
book = openpyxl.load_workbook("C:\\Users\\XT23495\\Downloads\\Python Selenium\\excel_data.xlsx")
sheet =book.active # to select the active sheet, control of sheet obj
Dict = {}
cell =sheet.cell(row=1, column=2)  # control is r 1 and c 2
print(cell.value)

sheet.cell(row=2, column=2).value = "rahul" #to write value to the sheet

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row) # to print the maximum rows

print(sheet.max_column) # to print maximum columns

print(sheet['A5'].value)

for i in range(1,sheet.max_row+1):  # to get rows
    if sheet.cell(row =i,column=1).value == "Testcase2":  # to add condition print only test cases 2 data ,next for loop only run if if condition is right

        #for i in range (1,sheet.max_row+1): print(sheet.cell(row=i, column =1).value)#coulm will 1 and row will iterate for all values

        for j in range(2,sheet.max_column+1):#to get columns
            #Dict["lastname"]="shetty
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value
            # row is 1 because header is in row 1 e.g Name firstname lastname in excel sheet

print(Dict)
